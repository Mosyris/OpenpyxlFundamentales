from openpyxl import Workbook
from openpyxl.chart import PieChart, ProjectedPieChart,Reference, PieChart3D
from openpyxl.chart.series import DataPoint
from copy import deepcopy
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import (GradientFillProperties, GradientStop, LinearShadeProperties)
from openpyxl.drawing.colors import SchemeColor

# create data structure for the file
data = [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40]
]
# activate the file
wb = Workbook()
ws = wb.active

# data won't be entered to the table until we append the rows
for row in data:
    ws.append(row)
pie = PieChart()
labels = Reference(ws,
                   min_col=1,
                   min_row=2,
                   max_row=5)
data = Reference(ws,
                 min_col=2,
                 min_row=1,
                 max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Pies sold by category"


# Cut the first slice out for the pie
slice = DataPoint(idx=0, explosion=20)
pie.series[0].data_points = [slice]
ws.add_chart(pie, "D1")

# in the same workbook, create another sheet called " Projection"
ws= wb.create_sheet(title="Projection")

data = [
    ['Page', 'Views'],
    ['Search', 95],
    ['Products', 4],
    ['Offers', 0.5],
    ['Sales', 0.5]
]
for row in data:
    ws.append(row)

projected_pie = ProjectedPieChart()
projected_pie.type = "pie"
projected_pie.splitType = "val" # split by value
labels = Reference(ws,
                   min_col=1,min_row=2,
                   max_row=5
)
data = Reference(ws,
                 min_col=2,
                 min_row=1,
                 max_row=5)
projected_pie.add_data(data, titles_from_data=True)
projected_pie.set_categories(labels)

ws.add_chart(projected_pie, "A10")

projected_bar = deepcopy(projected_pie)
projected_bar.type = "bar"
projected_bar.splitType = 'pos' # split by position
ws.add_chart(projected_bar, "A27")


ws= wb.create_sheet(title="3DPieCharts")

data = [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40]
]
for row in data:
    ws.append(row)

pie = PieChart3D()
labels = Reference(ws,
                   min_col=1,
                   min_row=2,
                   max_row=5)
data = Reference(ws,
                 min_col=2,
                 min_row=1,
                 max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Pies sold by category"
ws.add_chart(pie, "D1")

# Code for gradients piecharts
ws = wb.create_sheet("GradientChart")
data = [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40],
]

for row in data:
    ws.append(row)

pie = PieChart()
labels = Reference(ws,
                   min_col=1,
                   min_row=2,
                   max_row=5)
data = Reference(ws,
                 min_col=2,
                 min_row=1,
                 max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Pies sold by category"

# Cut the first slice out of the pie and apply a gradient to it
slice = DataPoint(
    idx=0,
    explosion=20,
    spPr=GraphicalProperties(
        gradFill=GradientFillProperties(
            gsLst=(
                GradientStop(
                    pos=0,
                    prstClr='blue'
                ),
                GradientStop(
                    pos=100000,
                    schemeClr=SchemeColor(
                        val='accent1',
                        lumMod=30000,
                        lumOff=70000
                    )
                )
            )
        )
    )
)
pie.series[0].data_points = [slice]

ws.add_chart(pie, "D1")

wb.save ("Pie.xlsx")